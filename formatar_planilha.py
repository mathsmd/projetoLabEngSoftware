import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def formatar_planilha(caminho="saida_final.xlsx"):
    # Reabrir o Excel pronto
    wb = load_workbook(caminho)
    ws = wb.active

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Cabeçalho estilizado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

        # Ajustar largura automaticamente
        max_length = max(len(str(cell.value)), *[
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(2, ws.max_row + 1)
        ])
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Exemplo: formatar colunas específicas como moeda
    colunas_moeda = ["VALOR DO PEDIDO", "VALOR COTAÇÃO CORREIOS", "VALOR COTAÇÃO JADLOG"]
    for col in colunas_moeda:
        if col in [c.value for c in ws[1]]:
            col_idx = [c.value for c in ws[1]].index(col) + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'

    wb.save(caminho)
    print(f"Planilha formatada com sucesso: {caminho}")
